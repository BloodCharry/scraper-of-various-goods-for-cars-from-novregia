import os
import time

import httpcore
import openpyxl
import requests
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

from googletrans import Translator

translator = Translator()

start = time.time()
base_url = 'https://www.autohifi.no'
url = 'https://www.autohifi.no/api/Menu/GetHtmlMenu' \
      '?nodeId=1003006&screensize=md&screensizePixels=992' \
      '&width=1920&height=1080&showMobileMenuCollapsed=false' \
      '&_=1679403050980'
html = requests.get(url)
soup = BeautifulSoup(html.text, 'lxml')
categories = soup.find_all('a')
max_categories = len(categories)

try:
    count_categories = int(
        input(f'Enter the number of categories you want to spar autohifi (max={max_categories}): ')
    )
except:
    print("Invalid input, please enter an integer:")
    count_categories = int()
    if count_categories < 1 or count_categories > max_categories \
            or count_categories is not int(count_categories):
        print('incorrect number of categories: you cannot specify less than zero, zero, or letters')
    while count_categories is not int(count_categories) or count_categories < 1:
        count_categories = int(
            input('Enter the number of categories you want to spar autohifi : '))


def autohifi_no():
    wb = openpyxl.Workbook()
    ws = wb.active
    header_row = (['Title', 'Category', 'Price', 'Article', 'Image', 'Description'])
    for col_num, header in enumerate(header_row, start=1):
        col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
        col_width = max(len(header) + 3, 10)  # минимальная ширина столбца - 10 символов
        col.width = col_width
    ws.append(header_row)
    for category in range(count_categories):
        category_url = categories[category].get('href')
        html = requests.get(f'https://www.autohifi.no{category_url}')
        soup = BeautifulSoup(html.text, 'lxml')
        links = soup.find_all('div', class_='WebPubElement pub-productlisting')
        if not links:
            print("category empty, try add more num of categories")
        for link in links:
            html = requests.get(base_url + link.find('a').get('href'))
            soup = BeautifulSoup(html.text, 'lxml')
            title = soup.find('div', class_="heading-container").text
            price = soup.find('span', class_="rrp-price-api").text[:-2]
            article = soup.find('span', class_="prd-num-label").text
            image = soup.find('img', class_='rsImg').get('src')
            description_nor = soup.find('div', class_="prod-text-content").text
            try:
                description_eng = translator.translate(
                    description_nor, src='no', dest='en'
                )
                data = {
                    'title': title,
                    'price': price,
                    'url': base_url + link.find('a').get('href'),
                    'category': base_url + category_url,
                    'image': base_url + image,
                    'article': article,
                    'description': description_eng.text
                }
                print(f'{base_url} - {data}')
                row = ([
                    data['title'], data['category'], data['price'],
                    data['article'], data['image'], data['description']
                ])
                for col_num, header in enumerate(row, start=1):
                    col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
                    col_width = max(len(header) + 3, 10)  # минимальная ширина столбца - 10 символов
                    col.width = col_width
                ws.append(row)
                script_dir = os.path.dirname(os.path.abspath(__file__))
                file_path = os.path.join(script_dir, "pars_result", "autohifi_no.xlsx")
                wb.save(file_path)
            except httpcore._exceptions.ReadTimeout:
                pass
            except TypeError:
                pass
            except IndexError:
                pass


autohifi_no()
print('data collection from autohifi ended')
print(f'time {time.time() - start} seconds')
