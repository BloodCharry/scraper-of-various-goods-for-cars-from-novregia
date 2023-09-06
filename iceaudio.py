import os
import time
from openpyxl.utils import get_column_letter
import httpcore
import openpyxl
import requests

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium import common
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options as ChromeOptions

from googletrans import Translator

options = ChromeOptions()
options.add_argument("--headless=new")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
translator = Translator()

start = time.time()
# count_categories = 1

base_url = 'https://iceaudio.no/'
base_url_for_item = 'https://iceaudio.no/index.php?func=varer&vnr='
r = requests.get(base_url)
soup = BeautifulSoup(r.text, 'lxml')
links_category = [
    base_url + i.find('a').get('href') for i in soup.find_all(
        'li') if '&ID' in i.find('a').get('href')
]

max_categories = len(links_category)

try:
    count_categories = int(
        input(f'Enter the number of categories you want to spar iceaudio (max={max_categories}): ')
    )
except:
    print("Invalid input, please enter an integer:")
    count_categories = int()
    if count_categories < 1 or count_categories > max_categories \
            or count_categories is not int(count_categories):
        print('incorrect number of categories: you cannot specify less than zero, zero, or letters')
    while count_categories is not int(count_categories) or count_categories < 1:
        count_categories = int(
            input('Enter the number of categories you want to spar iceaudio : '))


def iceaudio_no():
    global picture_2, picture_3, picture_4 \
        # picture_5, picture_6, picture_7, \
    # picture_8, picture_9, picture_10, picture_11, picture_12, picture_13
    wb = openpyxl.Workbook()
    ws = wb.active
    """
    Product ID, Main category,
    Category 1, Category 2, Category 3, Category 4, Category 5, 
    Brand, Product name, Product introduction, Car brand, Car model, Car year, 
    Main price, Discounted price, Product description, Picture 1, Picture 2, 
    Picture 3, Picture 4, Picture 5, Picture 6, Picture 7, Picture 8, 
    Picture 9, Picture 10, Picture 11, Picture 12, Picture 13.
    """
    # ws.append(['Title', 'Image', 'Description'])
    header_row = ([
        'Product ID', 'Product name', 'Main price', 'Product description',
        'Picture 1', 'Picture 2', 'Picture 2', 'Picture 4', 'Picture 5',
        'Picture 6', 'Picture 7', 'Picture 8', 'Picture 9', 'Picture 10',
        'Picture 11', 'Picture 12', 'Picture 13'
    ])
    for col_num, header in enumerate(header_row, start=1):
        col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
        col_width = max(len(header) + 3, 10)  # минимальная ширина столбца - 10 символов
        col.width = col_width
    ws.append(header_row)
    items = []
    for link in range(count_categories):
        try:
            r = requests.get(links_category[link])
        except:
            print('ConnectionResetError: 10054')
        soup = BeautifulSoup(r.text, 'lxml')
        for item in soup.find_all('input', {'name': 'varenr'}):
            items.append(item)
            product_id = item.get('value')
            driver.get(f"{base_url_for_item}{product_id}")

            title = driver.find_element(By.ID, 'PInfo_Top').text
            # image = driver.find_element(By.ID, 'PInfo_Left').find_element(
            #     By.TAG_NAME, 'img').get_attribute('src')
            # description = driver.find_element(By.ID, "PInfo_Right").text
            try:
                description = driver.find_element(
                    By.XPATH, '// *[ @ id = "PInfo_Right"] / ul').text
            except:
                description = 'description missing from page'
            price = driver.find_element(
                By.XPATH, '//*[@id="PInfo_Right"]/table/tbody/tr[3]/td[2]'
            ).text
            try:
                picture_1 = driver.find_element(
                    By.XPATH, '//*[@id="PInfo_Left"]/img'
                ).get_attribute('src')
            except:
                picture_1 = 'image missing'
            try:
                # //*[@id="PInfo_Left_bilder"]/a[1] #
                # //*[@id="PInfo_Left_bilder"]/a[1]/img
                picture_2 = driver.find_element(
                    By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[2]/img'
                ).get_attribute('src')
                picture_3 = driver.find_element(
                    By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[3]/img'
                ).get_attribute('src')
                picture_4 = driver.find_element(
                    By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[4]/img'
                ).get_attribute('src')
                # picture_5 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[5]/img'
                # ).get_attribute('src')
                # picture_6 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[6]/img'
                # ).get_attribute('src')
                # picture_7 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[7]/img'
                # ).get_attribute('src')
                # picture_8 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[8]/img'
                # ).get_attribute('src')
                # picture_9 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[9]/img'
                # ).get_attribute('src')
                # picture_10 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[10]/img'
                # ).get_attribute('src')
                # picture_11 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[11]/img'
                # ).get_attribute('src')
                # picture_12 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[12]/img'
                # ).get_attribute('src')
                # picture_13 = driver.find_element(
                #     By.XPATH, '//*[@id="PInfo_Left_bilder"]/a[13]/img'
                # ).get_attribute('src')
            except common.exceptions.NoSuchElementException:
                pass
            try:
                translation = translator.translate(
                    description, src='no', dest='en'
                )

                data = {'id': product_id,
                        'title': title,
                        'price': price,
                        'description': translation.text,
                        'picture_1': picture_1,
                        'picture_2': picture_2,
                        'picture_3': picture_3,
                        'picture_4': picture_4,
                        # 'picture_5': picture_5,
                        # 'picture_6': picture_6,
                        # 'picture_7': picture_7,
                        # 'picture_8': picture_8,
                        # 'picture_9': picture_9,
                        # 'picture_10': picture_10,
                        # 'picture_11': picture_11,
                        # 'picture_12': picture_12,
                        # 'picture_13': picture_13
                        }  # 'image': image,
                print(f'{base_url} - {data}')

                row = ([
                    data['id'], data['title'], data['price'],
                    data['description'], data['picture_1'], data['picture_2'],
                    data['picture_3'], data['picture_4'],  # data['picture_5'],
                    # data['picture_6'], data['picture_7'], data['picture_8'],
                    # data['picture_9'], data['picture_10'], data['picture_11'],
                    # data['picture_12'], data['picture_13']
                ])
                for col_num, header in enumerate(row, start=1):
                    col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
                    col_width = max(len(header) + 3, 10)
                    col.width = col_width
                ws.append(row)
                script_dir = os.path.dirname(os.path.abspath(__file__))
                file_path = os.path.join(script_dir, "pars_result", "iceaudio_no.xlsx")
                wb.save(file_path)
            except requests.exceptions.ConnectionError:
                pass
            except httpcore._exceptions.ReadTimeout:
                pass
            except TypeError:
                pass
            except IndexError:
                pass
            except ValueError:
                pass


iceaudio_no()
print('data collection from iceaudio ended')
print(f'time {time.time() - start} seconds')
