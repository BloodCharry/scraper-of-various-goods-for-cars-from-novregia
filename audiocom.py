import os
import time
import openpyxl
import requests
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

start = time.time()
base_url = 'https://www.audiocom.no'
menu_url = 'https://www.audiocom.no/api/Menu/GetHtmlMenu' \
           '?nodeId=2006202&screensize=sm&screensizePixels=768' \
           '&width=1920&height=1080&showMobileMenuCollapsed=false' \
           '&_=1679509214435'
menu_response = requests.get(menu_url)

soup = BeautifulSoup(menu_response.text, 'html.parser')
links_categories = [
    i.get('href') for i in soup.find_all('a') if i.get('href')[0] == '/'
]
max_categories = len(links_categories)
try:
    count_categories = int(
        input(f'Enter the number of categories you want to spar audiocom (max={max_categories}): ')
    )
except:
    print("Invalid input, please enter an integer:")
    count_categories = int()
    if count_categories < 1 or count_categories > max_categories \
            or count_categories is not int(count_categories):
        print('incorrect number of categories: you cannot specify less than zero, zero, or letters')
    while count_categories is not int(count_categories) or count_categories < 1:
        count_categories = int(
            input('Enter the number of categories you want to spar audiocom : '))



def audiocom_no():
    wb = openpyxl.Workbook()
    ws = wb.active
    header_row = (['Title', 'Item_id', 'Price', 'Image', 'Category', 'Url'])
    for col_num, header in enumerate(header_row, start=1):
        col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
        col_width = max(len(header) + 3, 10)  # минимальная ширина столбца - 10 символов
        col.width = col_width
    ws.append(header_row)
    for link in range(count_categories):
        r = requests.get(base_url + links_categories[link]).text
        soup = BeautifulSoup(r, 'lxml')
        items = soup.find_all('div', class_="AddHeaderContainer")
        for item in items:
            r = requests.get(base_url + item.find('a').get('href')).text
            soup = BeautifulSoup(r, 'lxml')
            price_elem = soup.find('span', class_='product-price-api')
            price = price_elem.text[:-2] if price_elem is not None else ''
            data = {
                'title': soup.find('div', class_='heading-container').text,
                'item_id': soup.find('div',
                                     class_='product-number-inner').text[12:],
                'price': price,
                'image': base_url + soup.find('img', class_='rsImg').get('src'),
                'category': base_url + links_categories[link],
                'url': base_url + item.find('a').get('href'),
            }
            print(f'{base_url} - {data}')
            row = ([
                data['title'], data['item_id'], data['price'],
                data['image'], data['category'], data['url']
            ])
            for col_num, header in enumerate(row, start=1):
                col = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
                col_width = max(len(header) + 3, 10)
                col.width = col_width
            ws.append(row)
            script_dir = os.path.dirname(os.path.abspath(__file__))
            file_path = os.path.join(script_dir, "pars_result", "audiocom_no.xlsx")
            wb.save(file_path)


audiocom_no()
print('data collection from audiocom ended')
print(f'time - {time.time() - start} seconds')
